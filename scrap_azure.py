from deep_translator import GoogleTranslator
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
import requests
import argostranslate.package
import argostranslate.translate
import pathlib


class AzureUpdatesScraper:
    def __init__(self, get_limit):
        self.get_url = 'https://azure.microsoft.com/ja-jp/updates/'
        self.get_limit = get_limit
        self.begin_url = 'https://azure.microsoft.com'
        self.translator = GoogleTranslator(source='auto', target='ja')
        self.full_urls = []
        self.pkg_infos = [
            self.ArgosTranslateInfo(from_code='en', to_code='ja'),
        ]
        # Packageインデクス情報をダウンロード確認 (アップデート確認)
        argostranslate.package.update_package_index()
        # 今回使用するパッケージ情報を取得し、パッケージをダウンロードし、そのパスを保存しておく
        # 2回目以降の実行では、pkg_infosのdownload_pathさえあれば、今回使用するパッケージを読み込みのみで実行可能
        for pkg_info in self.pkg_infos:
            if pkg_info.download_path is None or not pkg_info.download_path.exists():
                argostranslate.package.update_package_index()
                self.do_download(pkg_info)
            argostranslate.package.install_from_path(pkg_info.download_path)


    class ArgosTranslateInfo:
        def __init__(self, from_code: str = 'en', to_code: str = 'jp', download_path: pathlib.Path = None):
            self.from_code: str = from_code
            self.to_code: str = to_code
            self.download_path = download_path

        def do_download(self, install_package: argostranslate.package.AvailablePackage) -> None:
            # 今回使用するパッケージをダウンロードし、そのパスを保存しておく
            self.set_download_path(install_package.download())

        def set_download_path(self, download_path: pathlib.Path) -> None:
            self.download_path = download_path


    def do_download(self, pkg_info: ArgosTranslateInfo) -> None:
        # 使用可能な翻訳パッケージ情報を取得
        available_pkgs: list = argostranslate.package.get_available_packages()

        # 今回使う翻訳パッケージ情報を取得
        install_package: argostranslate.package.AvailablePackage = next(
            filter(
                lambda x: x.from_code == pkg_info.from_code and x.to_code == pkg_info.to_code,
                available_pkgs
            )
        )

        # ダウンロード実施 (パスも保存される)
        pkg_info.do_download(install_package)


    def do_translate(self, text: str, from_code: str = 'en', to_code: str = 'ja') -> str:
        translatedText: str = argostranslate.translate.translate(
            text, from_code, to_code)
        if to_code == 'ja':
            print('日本語に翻訳完了')
        return translatedText


    def scrape_url(self):
        try:
            response = requests.get(self.get_url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            urls = []
            for link in soup.find_all('a', {'data-test-element': 'update-entry-link'}):
                if len(urls) < self.get_limit:
                    urls.append(link.get('href'))
                else:
                    break
            urls.reverse()
            self.full_urls = [self.begin_url+url for url in urls]
            print(self.full_urls)
        except requests.exceptions.RequestException as e:
            print(f"要求失敗: {e}")
        except Exception as e:
            print(f"エラーが発生しました。: {e}")

    def excel_write(self):
        try:
            excel_file_path = 'test.xlsx'
            wb = load_workbook(excel_file_path)
            ws = wb.active
            for start_cell in range(3, ws.max_row+2):
                if ws[f'A{start_cell}'].value is None:
                    print(f'Excelの{start_cell}行目から書き込みます')
                    for i, url in enumerate(self.full_urls, start_cell):
                        response = requests.get(url)
                        soup = BeautifulSoup(response.text, 'html.parser')
                        # 題名取得
                        h1_tag = soup.find('h1')
                        # 日付取得
                        date_div_tag = soup.find('div', class_='row row-size3 column')
                        date_h6_tag = date_div_tag.find('h6')
                        date = date_h6_tag.text.split(':')[-1].strip()
                        dt = datetime.strptime(date, '%m月 %d, %Y')
                        date_tag = dt.strftime('%Y/%m/%d')
                        # タグの取得
                        tags_ul = soup.find('ul', class_='tags')
                        tags_li = tags_ul.find_all('li')
                        tags = [li.text for li in tags_li]
                        tags_str = '\n'.join(tags)
                        # 本文の取得
                        main_column_tags = soup.find('div', class_='column small-12')
                        main_tags = main_column_tags.find('div', class_='row column')
                        main_tag = main_tags.text.strip()
                        # 本文の日本語翻訳
                        main_tag_ja = self.do_translate(main_tag)
                        # ステータスの取得
                        sta_tag = soup.find('span', class_='status-indicator__label')
                        # セルに書き込み
                        if h1_tag and date_tag and tags_str and main_tag and main_tag_ja:
                            ws[f'A{i}'] = h1_tag.text
                            ws[f'A{i}'].hyperlink = url
                            ws[f'A{i}'].font = Font(color=Color("0563C1"), underline="single")
                            ws[f'B{i}'] = date_tag
                            if sta_tag is not None:
                                ws[f'C{i}'] = sta_tag.text
                            else:
                                print("タグが存在しません。")
                            ws[f'D{i}'] = tags_str
                            ws[f'E{i}'] = main_tag
                            ws[f'F{i}'] = main_tag_ja
                        else:
                            print('なにかの値がとれていません。')
                    break
            wb.save(excel_file_path)
            wb.close()
            return print('取得した値をExcelに書き込み完了')
        except Exception as e:
            print(f"エラーが発生しました。: {e}")

# 実行
get_limit_input = int(input('取得したい数を入力してください。：  '))
scraper_obj = AzureUpdatesScraper(get_limit_input)
scraper_obj.scrape_url()
scraper_obj.excel_write()